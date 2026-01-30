from typing import Union
import scipy.stats as sps
import scipy.optimize as spo
import numpy as np
from .utilities import _print_if
from openpyxl import load_workbook

from typing import Sequence, Tuple, Union
NumberArray = Union[np.ndarray, Sequence[float]]

class DustDistribution:
    """
    Representation of a mixture of Gaussian distributions in log10(diameter) space.

    This class fits a mixture of normals to cumulative distribution data
    and provides conversions between number, mass, and area distributed forms.

    Args:
        params: Flattened list/array of parameters `[w0,…,wN−1, mu0,…,muN−1, sig0,…,sigN−1]`
        type: One of 'mass', 'number', 'area' indicating distribution type
    """

    def __init__(self,
                 params: Union[None, NumberArray] = None,
                 type: Union[None, str] = None) -> None:

        self.n_components = None
        self.sub_dists = []
        self.weights = []
        self.type = []
        self.units = []
        if params is not None:
            assert type.lower() in [
                "mass",
                "number",
                "area",
            ], "Please supply a type (mass, number, area)"
            N = len(params) / 3
            assert (
                np.abs(N - np.floor(N)) < np.finfo(float).eps
            ), "Please specify parameters of each component as a 1D numpy.array([weights,mu,sigma])."

            N = int(np.floor(N))
            w, mu, sig = params[0:N], params[N : 2 * N], params[2 * N : :]
            self.n_components = N
            self.sub_dists = [sps.norm(loc=mu[ii], scale=sig[ii]) for ii in range(N)]
            self.weights = w
            self.type = type
            self._set_units()

    def pdf(self, x: NumberArray) -> np.ndarray:
        """
        Compute mixture PDF at `x` (log10 diameter).
        """
        pdf = 0
        for ii in range(self.n_components):
            pdf += self.weights[ii] * self.sub_dists[ii].pdf(x)
        return pdf

    def cdf(self, x: NumberArray) -> np.ndarray:
        """
        Compute mixture CDF at `x` (log10 diameter).
        """
        cdf = 0
        for ii in range(self.n_components):
            cdf += self.weights[ii] * self.sub_dists[ii].cdf(x)
        return cdf

    def mean(self) -> float:
        """
        Compute the weighted mean (log10 space) of the mixture.
        """
        m = 0
        sum_weights = sum(self.weights)
        for ii in range(self.n_components):
            m += self.weights[ii] / sum_weights * self.sub_dists[ii].mean()
        return m

    def icdf(self, p: float) -> float:
        """
        Inverse CDF via root finding on the mixture CDF.

        Args:
            p: probability in [0,1]
        Returns:
            x such that CDF(x) == p
        """

        def residual(x: float) -> float:
            return float(self.cdf(x) - p)

        result, = spo.fsolve(residual, self.mean())
        return result

    def _sse(self,
        params: NumberArray,
        log_diameter_values: NumberArray,
        pm_values: NumberArray) -> float:
        """Internal least-squares objective on cumulative values."""

        N = len(params) / 3
        assert (
            np.abs(N - np.floor(N)) < np.finfo(float).eps
        ), "Please specify parameters of each component as a 1D numpy.array([weights,mu,sigma])."
        N = int(np.floor(N))
        w, mu, sig = params[0:N], params[N : 2 * N], params[2 * N : :]
        self.n_components = N
        self.sub_dists = [sps.norm(loc=mu[ii], scale=sig[ii]) for ii in range(N)]
        self.weights = w

        return np.sum((self.cdf(log_diameter_values) - pm_values) ** 2)

    def fit(self,
        params0: NumberArray,
        log_diameter_values: NumberArray,
        cumulative_values: NumberArray,
        values_type: str = "mass",
        tol: float = 1e-3) -> spo.OptimizeResult:
        """
        Fit a mixture to data using bound-constrained least squares.

        Args:
            params0: initial guess parameters [w,mu,sigma]
            log_diameter_values: sorted log10 diameter values
            cumulative_values: empirical cumulative values at those points
            values_type: distribution type of fitted data
        Returns:
            The SciPy OptimizeResult
        """

        N = len(params0) / 3
        assert (
            np.abs(N - np.floor(N)) < np.finfo(float).eps
        ), "Please specify parameters of each component as a 1D numpy.array([weights,mu,sigma])."
        N = int(np.floor(N))

        def fun(x):
            return self._sse(x, log_diameter_values, cumulative_values)

        # construct bounds
        lower_bound_w = [0] * N
        lower_bound_mu = [-np.inf] * N
        lower_bound_sig = [0 + tol] * N
        lb = lower_bound_w + lower_bound_mu + lower_bound_sig  # join lists
        ub = [np.inf] * len(lb)

        bnds = spo.Bounds(lb=lb, ub=ub, keep_feasible=True)
        res = spo.minimize(fun, params0, bounds=bnds, tol=1e-8)

        params = res.x
        N = int(np.floor(len(params) / 3))
        w, mu, sig = params[0:N], params[N : 2 * N], params[2 * N : :]
        self.n_components = N
        self.sub_dists = [sps.norm(loc=mu[ii], scale=sig[ii]) for ii in range(N)]
        self.weights = w
        self.type = values_type
        self._set_units()

        return res

    def _set_units(self)-> None:
        """Internal: set the `units` string based on distribution type."""

        if self.type.lower() == "mass":
            self.units = r"$\frac{\mu g \cdot m^{{-3}}}{d(\log(D))}$"
        elif self.type.lower() == "number":
            self.units = r"$\frac{cm^{{-3}} ] }{d(\log(D))}$"
        elif self.type.lower() == "area":
            self.units = r"$\frac{ m^2 \cdot m^{{-3}}}{d(\log(D))}$"

    def convert_to_number(self, rho: float = None) -> None:
        """
        Convert this distribution to number.

        Args:
            rho: particle density in correct units (required for mass → number)
        """
        if self.type.lower() == "number":
            print("Type is already " "number" " ")
        elif self.type.lower() == "mass":
            assert isinstance(rho, float), "Particle density must be a scalar float."
            new_subs = []
            new_weights = []
            for ii in range(self.n_components):
                Mi = self.weights[ii]
                mbari = self.sub_dists[ii].mean()
                si = self.sub_dists[ii].std()

                mi = mbari - 3 * si**2 / np.log10(np.e)
                b = 2 * mi + 6 * si**2 / np.log10(np.e)
                Ni = 6 * Mi / np.pi / rho * np.exp((mi**2 - 0.25 * b**2) / 2 / si**2)

                new_weights.append(Ni * 1e3)
                ns = sps.norm(loc=mi, scale=si)
                new_subs.append(ns)

            self.weights = new_weights
            self.sub_dists = new_subs
            self.type = "number"
            self._set_units()
        elif self.type.lower() == "area":
            new_subs = []
            new_weights = []
            for ii in range(self.n_components):
                Ai = self.weights[ii]
                mbari = self.sub_dists[ii].mean()
                si = self.sub_dists[ii].std()

                mi = mbari - 2 * si**2 / np.log10(np.e)
                Ni = (
                    Ai
                    / np.pi
                    * 4
                    * np.exp((mi**2 - (mi - si**2 / np.log10(np.e)) ** 2) / 2 / si**2)
                )

                new_weights.append(Ni * 1e6)
                ns = sps.norm(loc=mi, scale=si)
                new_subs.append(ns)

            self.weights = new_weights
            self.sub_dists = new_subs
            self.type = "number"
            self._set_units()

    def convert_to_mass(self, rho):
        """
        Convert this distribution to mass.

        Args:
            rho: particle density in correct units (required for number → mass)
        """
        assert isinstance(rho, float), "Particle density must be a scalar float."
        if self.type.lower() == "mass":
            print("Type is already " "mass" " ")
        elif self.type.lower() == "number":
            new_subs = []
            new_weights = []
            for ii in range(self.n_components):
                Ni = self.weights[ii]
                mi = self.sub_dists[ii].mean()
                si = self.sub_dists[ii].std()

                b = 2 * mi + 6 * si**2 / np.log10(np.e)
                mbari = b / 2.0
                Mi = Ni * np.pi * rho / 6 * np.exp(-(mi**2 - 0.25 * b**2) / 2 / si**2)

                new_weights.append(Mi * 1e-3)
                ns = sps.norm(loc=mbari, scale=si)
                new_subs.append(ns)

            self.weights = new_weights
            self.sub_dists = new_subs
            self.type = "mass"
            self._set_units()
        elif self.type.lower() == "area":
            print("Convert to number first. ")

    def convert_to_area(self, rho=None):
        """
        Convert distribution to area.

        Args:
            rho: particle density in correct units (required for number <-> mass)
        """

        if self.type.lower() == "area":
            print("Type is already " "area" " ")
        elif self.type.lower() == "number":
            new_subs = []
            new_weights = []
            for ii in range(self.n_components):
                Ni = self.weights[ii]
                mi = self.sub_dists[ii].mean()
                si = self.sub_dists[ii].std()

                mbari = mi + 2 * si**2 / np.log10(np.e)
                Ai = (
                    Ni
                    * np.pi
                    / 4
                    * np.exp(-(mi**2 - (mi + si**2 / np.log10(np.e)) ** 2) / 2 / si**2)
                )

                new_weights.append(Ai * 1e-6)
                ns = sps.norm(loc=mbari, scale=si)
                new_subs.append(ns)

            self.weights = new_weights
            self.sub_dists = new_subs
            self.type = "area"
            self._set_units()
        elif self.type.lower() == "mass":
            if rho is None:
                raise ValueError("Rho cannot be None to convert from mass")

            assert isinstance(rho, float), "Particle density must be a scalar float."
            self.convert_to_number(rho)
            self.convert_to_area()

    def write_to_file(self,
        file_name: str,
        sheet_name: str,
        kind: str = "number",
        rho: float = None,
        verbose: bool = True) -> None:
        """
        Write the fitted distribution to an Excel file.

        Args:
            file_name: path to .xlsx
            sheet_name: sheet containing dust parameters
            kind: 'number', 'mass', or 'area'
            rho: density required for conversions
            verbose: whether to print status
        """
        _print_if("Writing dust distribution to file " + file_name, verbose)

        # ensure kind is correct
        if kind.lower() == "number":
            self.convert_to_number(rho)
        elif kind.lower(rho) == "mass":
            self.convert_to_mass(rho)
        elif kind.lower(rho) == "area":
            self.convert_to_area(rho)
        else:
            raise ValueError("kind not recognized.")

        # convert to strings and join with ";" delimiter
        weight_str = [str(s) + ";" for s in self.weights]
        mu_str = [str(10 ** (self.sub_dists[ii].mean())) for ii in range(self.n_components)]
        sig_str = [str(10 ** (self.sub_dists[ii].std())) for ii in range(self.n_components)]
        weight_str = "".join(weight_str)[0:-1]
        mu_str = "".join(mu_str)[0:-1]
        sig_str = "".join(sig_str)[0:-1]

        # write data
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
        for cell in ws["A"]:
            if cell.value == "N_size":
                ws.cell(row=cell.row, column=2).value = self.n_components
                ws.cell(row=cell.row, column=4).value = ""
            elif cell.value == "Nd":
                ws.cell(row=cell.row, column=2).value = weight_str
                ws.cell(row=cell.row, column=4).value = ""
            elif cell.value == "mu":
                ws.cell(row=cell.row, column=2).value = mu_str
                ws.cell(row=cell.row, column=4).value = ""
            elif cell.value == "sigma":
                ws.cell(row=cell.row, column=2).value = sig_str
                ws.cell(row=cell.row, column=4).value = ""

        wb.save(filename=file_name)
        wb.close()
