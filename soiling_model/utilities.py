import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import scipy.stats as sps
import scipy.optimize as spo
import numpy as np
from copy import deepcopy
from openpyxl import load_workbook
import os
import miepython
from collections import defaultdict

def _print_if(s,verbose):
    # Helper function to control level of output display.
    if verbose:
        print(s)

def _ensure_list(s):
    if not isinstance(s,list):
        s = [s]
    return s

def _check_keys(simulation_data,reflectance_data):
    for ii in range(len(simulation_data.time.keys())):
            if simulation_data.file_name[ii] != reflectance_data.file_name[ii]:
                raise ValueError("Filenames in simulation data and reflectance do not match. Please ensure you imported the same list of files for both.")

def _import_option_helper(file_list,option):
    if isinstance(option,(list,np.ndarray)):
        assert len(file_list) == len(option), "Please supply a list for {option} containing one string for each experiment. Or, supply a single global type by specifying a string. "
    else:
        option = [option]*len(file_list)
    
    return option

def simple_annual_cleaning_schedule(n_sectors,n_trucks,n_cleans,dt=1,n_sectors_per_truck=1):
    T_days = 365
    n_hours = int(T_days*(24/dt)) # number of hours between cleanings
    clean_interval = np.floor(T_days/n_cleans)
    min_clean_interval = np.ceil(n_sectors/n_trucks/n_sectors_per_truck)
    if clean_interval < min_clean_interval:
        clean_interval = min_clean_interval
        n_cleans = int(np.floor(T_days/clean_interval))
        print("Warning: Cannot clean that many times. Setting number of cleans = "+str(n_cleans))

    # evenly space cleaning ends
    clean_ends = np.linspace(0,n_hours-1,num=n_cleans+1,dtype=int)
    clean_ends = np.delete(clean_ends,-1) # remove the last clean since (clean at 0 takes care of this) 
    
    # shift schedule
    cleans = np.zeros((n_sectors,n_hours))
    for ii in range(n_trucks*n_sectors_per_truck,n_sectors,n_trucks*n_sectors_per_truck):
        idx0 = n_sectors-ii
        idx1 = n_sectors-(ii-n_trucks*n_sectors_per_truck)
        idx_col = clean_ends-(24/dt)*(int(ii/n_trucks/n_sectors_per_truck)-1)
        for jj in idx_col.astype(int):
            if jj<0:
                cc = jj + 365*24
                cleans[idx0:idx1,cc] = 1
            else:
                cleans[idx0:idx1,jj] = 1

    # take care of remainder (first day of a field clean)
    if idx0 != 0:
        idx_col = clean_ends-(24/dt)*int(ii/n_trucks/n_sectors_per_truck)
        for jj in idx_col.astype(int):
            if jj<0:
                cc = jj + 365*24
                cleans[0:idx0,cc] = 1
            else:
                cleans[0:idx0,jj] = 1
    return cleans

def plot_experiment_data(simulation_inputs,reflectance_data,experiment_index,lgd_label=None,figsize=(7,12),lgd_size = 15):
    sim_data = simulation_inputs
    reflect_data = reflectance_data
    f = experiment_index

    fig,ax = plt.subplots(nrows=5,sharex=True,figsize=figsize)
    # fmt = r"${0:s}^\circ$"
    # fmt = r'${0:s}$'
    ave = reflect_data.average[f]
    t = reflect_data.times[f]
    std = reflect_data.sigma[f]
    if lgd_label == None:
        names = ["M"+str(ii+1) for ii in range(ave.shape[1])]
    else:
        names = lgd_label
    # for ii in range(ave.shape[1]):
    #     # remove NaNs
    #     ax[0].errorbar(t,ave[:,ii],yerr=1.96*std[:,ii],label=(names[ii][0:5]),marker='o',capsize=4.0)
    for ii in range(ave.shape[1]):
        # remove NaNs for irregular data measurements
        valid_indices = ~np.isnan(ave[:,ii])
        t_valid = t[valid_indices].squeeze()
        ave_valid = ave[valid_indices,ii].squeeze()
        std_valid = std[valid_indices,ii].squeeze()
        ax[0].errorbar(t_valid,ave_valid,yerr=1.96*std_valid,label=(names[ii][0:5]),marker='o',capsize=4.0)

    ax[0].grid(True) 
    label_str = r"Reflectance at {0:.1f} $^{{\circ}}$".format(reflect_data.reflectometer_incidence_angle[f]) 
    ax[0].set_ylabel(label_str)
    ax[0].legend(fontsize=lgd_size,ncol=len(ave)//2)

    ax[1].plot(sim_data.time[f],sim_data.dust_concentration[f],color='brown',label="measurements")
    # ax[1].plot(sim_data.hours[f],sim_data.hourly_dust_avg[f],color='black',ls='--',label="hourly")
    ax[1].plot(sim_data.days[f],sim_data.daily_dust_avg[f],color='black',ls='--',label="daily")
    ax[1].axhline(y=np.nanmean(sim_data.dust_concentration[f]),color='brown',ls='--',label = r"Average = {0:.2f}".format(np.nanmean(sim_data.dust_concentration[f])))
    label_str = r'{0:s} [$\mu g\,/\,m^3$]'.format(sim_data.dust_type[0])
    ax[1].set_ylabel(label_str,color='brown',fontsize=20)
    ax[1].tick_params(axis='y', labelcolor='brown')
    YL_dust = 3*np.nanmean(sim_data.dust_concentration[f])
    ax[1].set_ylim((0,YL_dust))
    ax[1].grid(True)
    ax[1].legend(fontsize=lgd_size)

    # Rain intensity, if available
    if len(sim_data.rain_intensity)>0: # rain intensity is not an empty dict
        ax[2].plot(sim_data.time[f],sim_data.rain_intensity[f])
    else:
        rain_nan = np.nan*np.ones(sim_data.time[f].shape)
        ax[2].plot(sim_data.time[f],rain_nan)
    
    ax[2].set_ylabel(r'Rain [mm/hour]',color='blue')
    ax[2].tick_params(axis='y', labelcolor='blue')
    YL = ax[2].get_ylim()
    ax[2].set_ylim((0,YL[1]))
    ax[2].grid(True)

    ax[3].plot(sim_data.time[f],sim_data.wind_speed[f],color='green',label="measurements")
    ax[3].axhline(y=np.nanmean(sim_data.wind_speed[f]),color='green',ls='--',label = r"Average = {0:.2f}".format(np.nanmean(sim_data.wind_speed[f])))
    label_str = r'Wind Speed [$m\,/\,s$]'
    ax[3].set_ylabel(label_str,color='green')
    ax[3].set_xlabel('Date')
    ax[3].tick_params(axis='y', labelcolor='green')
    ax[3].grid(True)
    ax[3].legend(fontsize=lgd_size)

    if len(sim_data.relative_humidity)>0: 
        ax[4].plot(sim_data.time[f],sim_data.relative_humidity[f],color='black',label="measurements")
        ax[4].axhline(y=np.nanmean(sim_data.relative_humidity[f]),color='black',ls='--',label = r"Average = {0:.2f}".format(np.nanmean(sim_data.relative_humidity[f])))
    else:
        rain_nan = np.nan*np.ones(sim_data.time[f].shape)
        ax[4].plot(sim_data.time[f],rain_nan)
    
    label_str = r'Relative Humidity [%]'
    ax[4].set_ylabel(label_str,color='black')
    ax[4].set_xlabel('Date')
    ax[4].tick_params(axis='y', labelcolor='black')
    ax[4].grid(True)
    ax[4].legend(fontsize=lgd_size)
    
    if len(sim_data.wind_direction)>0: 
        figwr,axwr = wind_rose(sim_data,f)
        figwr.tight_layout()

    fig.autofmt_xdate()
    fig.tight_layout()

    return fig,ax

def plot_experiment_PA(simulation_inputs,reflectance_data,experiment_index,figsize=(7,12)):
    sim_data = simulation_inputs
    reflect_data = reflectance_data
    f = experiment_index

    fig,ax = plt.subplots(nrows=3,sharex=True,figsize=figsize)
    # fmt = r"${0:s}^\circ$"
    fmt = "${0:s}$"
    ave = reflect_data.average[f]
    t = reflect_data.times[f]
    std = reflect_data.sigma[f]
    # names = ["M"+str(ii+1) for ii in range(ave.shape[1])]
    names = ["SE1",	"SE2",	"SE3",	"SE4",	"SE5",	"NW1",	"NW2",	"NW3",	"NW4",	"NW5"]

    for ii in range(ave.shape[1]):
        ax[0].errorbar(t,ave[:,ii],yerr=1.96*std[:,ii],label=fmt.format(names[ii]),marker='o',capsize=4.0)

    ax[0].grid(True) 
    label_str = r"Reflectance at {0:.1f} $^{{\circ}}$".format(reflect_data.reflectometer_incidence_angle[f]) 
    ax[0].set_ylabel(label_str)
    ax[0].legend(loc='upper left', bbox_to_anchor=(1, 1))
    # ax[0].set_ylim(0.86, 0.97)

    ax[1].plot(sim_data.time[f],sim_data.dust_concentration[f],color='brown',label="Measurements")
    ax[1].axhline(y=sim_data.dust_concentration[f].mean(),color='brown',ls='--',label = "Average")
    label_str = r'{0:s} [$\mu g\,/\,m^3$]'.format(sim_data.dust_type[0])
    ax[1].set_ylabel(label_str,color='brown')
    ax[1].tick_params(axis='y', labelcolor='brown')
    ax[1].grid(True)
    ax[1].legend()
    title_str = sim_data.dust_type[f] + r" (mean = {0:.2f} $\mu g$/$m^3$)" 
    ax[1].set_title(title_str.format(sim_data.dust_concentration[f].mean()),fontsize=10)
    # ax[1].set_ylim(0,70)

    # # Rain intensity, if available
    # if len(sim_data.rain_intensity)>0: # rain intensity is not an empty dict
    #     ax[2].plot(sim_data.time[f],sim_data.rain_intensity[f])
    # else:
    #     rain_nan = np.nan*np.ones(sim_data.time[f].shape)
    #     ax[2].plot(sim_data.time[f],rain_nan)
    
    # ax[2].set_ylabel(r'Rain [mm/hour]',color='blue')
    # ax[2].tick_params(axis='y', labelcolor='blue')
    # YL = ax[2].get_ylim()
    # ax[2].set_ylim((0,YL[1]))
    # ax[2].grid(True)

    ax[2].plot(sim_data.time[f],sim_data.wind_speed[f],color='green',label="Measurements")
    ax[2].axhline(y=sim_data.wind_speed[f].mean(),color='green',ls='--',label = "Average")
    label_str = r'Wind Speed [$m\,/\,s$]'
    ax[2].set_ylabel(label_str,color='green')
    ax[2].set_xlabel('Date')
    ax[2].tick_params(axis='y', labelcolor='green')
    ax[2].grid(True)
    ax[2].legend()
    title_str = "Wind Speed (mean = {0:.2f} m/s)".format(sim_data.wind_speed[f].mean())
    ax[2].set_title(title_str,fontsize=10)
    # ax[2].set_ylim(0,8)

    # if len(sim_data.relative_humidity)>0: 
    #     ax[4].plot(sim_data.time[f],sim_data.relative_humidity[f],color='black',label="measurements")
    #     ax[4].axhline(y=sim_data.relative_humidity[f].mean(),color='black',ls='--',label = "Average")
    # else:
    #     rain_nan = np.nan*np.ones(sim_data.time[f].shape)
    #     ax[4].plot(sim_data.time[f],rain_nan)
    
    # label_str = r'Relative Humidity [%]'
    # ax[4].set_ylabel(label_str,color='black')
    # ax[4].set_xlabel('Date')
    # ax[4].tick_params(axis='y', labelcolor='black')
    # ax[4].grid(True)
    # ax[4].legend()
    
    fig.autofmt_xdate()
    fig.tight_layout()

    return fig,ax

def trim_experiment_data(simulation_inputs,reflectance_data,trim_ranges):
    sim_dat = deepcopy(simulation_inputs)
    ref_dat = deepcopy(reflectance_data)
    files = sim_dat.time.keys()

    # Ensure sim_dat has hourly_dust_avg and daily_dust_avg attributes
    if not hasattr(sim_dat, 'hourly_dust_avg'):
        sim_dat.hourly_dust_avg = {}
        sim_dat.hours = {}
    if not hasattr(sim_dat, 'daily_dust_avg'):
        sim_dat.daily_dust_avg = {}
        sim_dat.days = {}

    for f in files:
        if isinstance(trim_ranges,list):  
            assert isinstance(trim_ranges[f],list) or isinstance(trim_ranges[f],np.ndarray), "trim_ranges must be a list of lists or a list of 1D np.arrays"
            lb = trim_ranges[f][0]
            ub = trim_ranges[f][1]
        elif trim_ranges=="reflectance_data":
            assert ref_dat is not None, "Reflectance data must be supplied for trim_ranges==""reflectance_data"""
            lb = ref_dat.times[f][0]
            ub = ref_dat.times[f][-1]
        elif trim_ranges == "simulation_inputs":
            lb = sim_dat.time[f].iloc[0]
            ub = sim_dat.time[f].iloc[-1]
        else:
            raise ValueError("""Value of trim_ranges not recognized. Must be a list of lists/np.array [lb,ub], """+\
                """ "reflectance_data" or "simulation_inputs" """)

        # trim simulation data
        mask = (sim_dat.time[f]>=lb.astype('datetime64[ns]')) & (sim_dat.time[f]<=ub.astype('datetime64[ns]'))   # .astype('datetime64[ns]') guarantees compatibility with Pandas Timestamp
        if all(mask==0):
            raise ValueError(f"Provided date range of {lb} to {ub} for file {sim_dat.file_name[f]} excludes all data.")
            
        sim_dat.time[f] = sim_dat.time[f][mask]
        sim_dat.time_diff[f] = sim_dat.time_diff[f][mask]
        sim_dat.air_temp[f] = sim_dat.air_temp[f][mask]
        sim_dat.wind_speed[f] = sim_dat.wind_speed[f][mask]
        sim_dat.dust_concentration[f] = sim_dat.dust_concentration[f][mask]
        if len(sim_dat.rain_intensity)>0:
            sim_dat.rain_intensity[f] = sim_dat.rain_intensity[f][mask]
        if len(sim_dat.dni)>0:
            sim_dat.dni[f] = sim_dat.dni[f][mask]
        if len(sim_dat.relative_humidity)>0:
            sim_dat.relative_humidity[f] = sim_dat.relative_humidity[f][mask]
        if len(sim_dat.wind_direction)>0:
            sim_dat.wind_direction[f] = sim_dat.wind_direction[f][mask]
        
        # Calculate hourly and daily averages of dust_concentration
        dust_conc_temp = pd.Series(sim_dat.dust_concentration[f], index=pd.to_datetime(sim_dat.time[f]))
        hourly_avg = dust_conc_temp.resample('h').mean()
        daily_avg = dust_conc_temp.resample('D').mean()

        sim_dat.hourly_dust_avg[f] = hourly_avg
        sim_dat.daily_dust_avg[f] = daily_avg            
        sim_dat.hours[f] = hourly_avg.index  # Store hourly timestamps
        sim_dat.days[f] = daily_avg.index    # Store daily timestamps

        if reflectance_data is not None:
            # trim reflectance data
            if len(ref_dat.tilts)>0:
                mask_indices = mask.index[mask].to_numpy()  # Extract the indices where mask is True and  convert to NumPy array
                ref_dat.tilts[f] = ref_dat.tilts[f][:, mask_indices]
                # ref_dat.tilts[f] = ref_dat.tilts[f][:,mask]
            mask_ref = (ref_dat.times[f]>=lb) & (ref_dat.times[f]<=ub)
            ref_dat.times[f] = ref_dat.times[f][mask_ref] 
            ref_dat.average[f] = ref_dat.average[f][mask_ref,:]
            ref_dat.delta_ref[f] = ref_dat.delta_ref[f][mask_ref,:]
            ref_dat.sigma[f] = ref_dat.sigma[f][mask_ref,:]
            ref_dat.sigma_of_the_mean[f] = ref_dat.sigma_of_the_mean[f][mask_ref,:]

            ref_dat.prediction_indices[f] = []
            ref_dat.prediction_times[f] = []
            time_grid = sim_dat.time[f]
            for m in ref_dat.times[f]:
                ref_dat.prediction_indices[f].append(np.argmin(np.abs(m.astype('datetime64[ns]')-time_grid)))        # .astype('datetime64[ns]') guarantees compatibility with Numpy datetime64
                # ref_dat.prediction_times[f].append(time_grid.iloc[ref_dat.prediction_indices[f]])
                # ref_dat.rho0[f] = np.nanmax(ref_dat.average[f], axis=0) # this now avoid issues in case the first value is a NaN (it may happen if a mirror or heliostat is added later)
            ref_dat.prediction_times[f].extend(time_grid.iloc[ref_dat.prediction_indices[f]].tolist())
            ref_dat.rho0[f] = np.nanmax(ref_dat.average[f], axis=0) # this now avoid issues in case the first value is a NaN (it may happen if a mirror or heliostat is added later)
            elapsed_time = (ref_dat.times[f][-1] - ref_dat.times[f][0]) / np.timedelta64(1, 'D')            # compute total time in days as a np.float64
            ref_dat.soiling_rate[f] = (ref_dat.average[f][0]-ref_dat.average[f][-1])/elapsed_time*100       # compute soiling rates in p.p./day for each mirror
            
        sim_dat.time[f] = sim_dat.time[f].reset_index(drop=True) # reset indices to start at 0 to allow repeated iterations
            
                 
    
    return sim_dat,ref_dat

def daily_average(ref_dat,time_grids,dt=None):

    # prediction indeces and times
    # tilts

    ref_dat_new = deepcopy(ref_dat)
    num_files = len(ref_dat.file_name)
    for f in range(num_files):
        num_mirrors = ref_dat.average[f].shape[1]
        df = pd.DataFrame({"times":ref_dat.times[f],
                           "day":ref_dat.times[f].astype('datetime64[D]')})
        times = df.groupby('day')['times'].mean().values
        # ref_dat_new.tilts[f] = [] # the averaging for the tilts needs to be done seperately
        if dt is None:
            ref_dat_new.times[f] = times
        else:
            ref_dat_new.times[f] = times.astype(f'datetime64[{int(dt[f])}s]')

        num_times = len(times)
        ref_dat_new.sigma[f] = np.zeros((num_times,num_mirrors))
        ref_dat_new.average[f] = np.zeros((num_times,num_mirrors))
        ref_dat_new.sigma_of_the_mean[f] = np.zeros((num_times,num_mirrors))
        ref_dat_new.delta_ref[f] = np.zeros((num_times,num_mirrors))
        ref_dat_new.soiling_rate[f] = np.zeros(ref_dat.rho0[f].shape)
        for ii in range(num_mirrors):
            df = pd.DataFrame( {"average":ref_dat.average[f][:,ii],
                                "sigma": ref_dat.sigma[f][:,ii],
                                "day":ref_dat.times[f].astype('datetime64[D]')
                                })
            
            daily = df.groupby("day")
            N = daily.count()['sigma'].values
            sum_var = df.groupby('day')['sigma'].apply(lambda x: sum(x.dropna()**2)).values
            ref_dat_new.sigma[f][:,ii] = np.sqrt(sum_var/N) # pooled variance
            
            ref_dat_new.sigma_of_the_mean[f][:,ii] = (ref_dat_new.sigma[f][:,ii] / 
                                np.sqrt(ref_dat_new.number_of_measurements[f]))

            ref_dat_new.average[f][:,ii] = daily.mean().average.values
            ref_dat_new.prediction_indices[f] = []
            ref_dat_new.prediction_times[f] = []
            
            # handle case when time_grids is a pandas something
            if isinstance(time_grids[f],(pd.Series,pd.DataFrame)):
                tg = time_grids[f].values
            else:
                tg = deepcopy(time_grids[f])

            for m in ref_dat_new.times[f]:
                idx = np.argmin(np.abs(m-tg))
                ref_dat_new.prediction_indices[f].append(idx)        
            ref_dat_new.prediction_times[f].append(tg[ref_dat_new.prediction_indices[f]])
        ref_dat_new.delta_ref[f] = np.vstack((np.zeros((1, ref_dat_new.average[f].shape[1])), -np.diff(ref_dat_new.average[f], axis=0)))  # compute reflectance loss between subsequent measurements (0 given to first timestamp)
        elapsed_time = (ref_dat_new.times[f][-1] - ref_dat_new.times[f][0]) / np.timedelta64(1, 'D')            # compute total time in days as a np.float64
        ref_dat_new.soiling_rate[f] = (ref_dat_new.average[f][0]-ref_dat_new.average[f][-1])/elapsed_time*100   # compute soiling rates in p.p./day for each mirror

    return ref_dat_new
            
def sample_simulation_inputs(historical_files,window=np.timedelta64(30,"D"),N_sample_years=10,\
                             sheet_name=None,output_file_format="sample_{0:d}.xlsx",\
                             dt=np.timedelta64(3600,'s'),verbose=True):

    # load in historical data files into a single pandas dataframe
    df = pd.DataFrame()
    for f in historical_files:
        fi = pd.read_excel(f,sheet_name=sheet_name)

        #check that time difference is equal to time grid
        if not np.all( fi['Time'].diff()[1::] == dt ): # omit first time, which is NaT
            raise ValueError("Time in file "+f+" is inconsistent with specified dt")

        fi['day'] = fi['Time'].apply(lambda x:x.day) 
        fi['month'] = fi['Time'].apply(lambda x:x.month)
        fi['year'] = fi['Time'].apply(lambda x:x.year)
        df = pd.concat((df,fi),ignore_index=True)

    # Create N_sample_years by sampling days from the historical dataset around a windwow
    t0 = pd.Timestamp( np.datetime64('now').astype('datetime64[Y]').astype('datetime64[m]') ) # t0 is the beginning of the current year
    tf = pd.Timestamp( t0 + np.timedelta64(365,'D') ) # t0 is the beginning of the current year

    dt_str = str(dt.astype('timedelta64[s]').astype('int'))+'s'
    time_grid = pd.date_range(start=t0,end=tf,freq=dt_str)
    day_grid = pd.date_range(start=t0,end=tf,freq="D")
    
    for n in range(N_sample_years):
        samples = pd.DataFrame(columns=df.columns)
        _print_if("Building sample {0:d} of {1:d}".format(n+1,N_sample_years),verbose)
        for ii in range(len(day_grid)-1):
            t = day_grid[ii]

            # samples days in the window
            sample_days = pd.date_range(start=t-window/2,end=t+window/2,freq="D")
            idx = np.random.randint(0,high=len(sample_days))
            sample_day = sample_days[idx]

            # select a random year of the selected day
            mask = (df.day==sample_day.day) & (df.month==sample_day.month)
            sample_years = np.unique(df.year[mask])
            idx_y = np.random.randint(0,high=len(sample_years))
            sample_year = sample_years[idx_y]

            # select 24-hour period that corresponds to the sampled dat in the historical database
            sample_mask = (df.year==sample_year) & mask

            samples = pd.concat((samples,df[sample_mask]),ignore_index=True)

        samples['Time'] = time_grid[0:-1]
        samples.set_index('Time',inplace=True)
        samples.drop(labels=['day','month','year'],axis=1,inplace=True)
        samples.to_excel(output_file_format.format(n),sheet_name=sheet_name)

def _extinction_function(diameters,lambdas,intensities,acceptance_angle,
                         refractive_index,grid_size_mu=int(1e4),
                         grid_size_x=1000,verbose=False):
    
    # theta_s = np.radians(np.linspace(-180,180,grid_size_theta_s)) # angle of scattering (\theta=0 is direction of radiation)
    m = refractive_index
    lam = lambdas/1000 # nm -> µm
    E = intensities*1000 #W/m^2/nm -> W/m^2/µm

    # set up grids
    # mu = np.sort(np.cos(theta_s)) 
    aa_cos = np.cos(acceptance_angle)
    mu = np.linspace(-1,aa_cos,num=grid_size_mu)
    _print_if(f"\t Acceptance angle cosine = {aa_cos:.6f}",verbose)

    # making lookup table in x
    min_x = np.pi*np.min(diameters)/np.max(lam)
    max_x = np.pi*np.max(diameters)/np.min(lam)
    xg = np.logspace(np.log10(min_x),np.log10(max_x),grid_size_x)
    Qxg = np.zeros(xg.shape)
    for ii,x in enumerate(xg):
        scat = miepython.i_unpolarized(m,x,mu,'qext')
        Qxg[ii] = np.trapz(scat,mu)
    
    # apply look up table to data
    Qx = np.zeros((len(diameters),len(lam)))
    for ii,d in enumerate(diameters):
        for jj,lamjj in enumerate(lam):
            x = np.pi*d/lamjj
            Qx[ii,jj] = np.interp(x,xg,Qxg)
    gamma = 2*np.pi*np.trapz(Qx*E,x=lam,axis=1) # for unit irradiance
    
    return gamma

def _same_ext_coeff(helios,simulation_data):
    
    sim_dat = simulation_data
    dust = sim_dat.dust
    D = dust.D
    refractive_index = dust.m
    lam = sim_dat.source_wavelength
    intensities = sim_dat.source_normalized_intensity
    phia = helios.acceptance_angles

    files = list(sim_dat.file_name.keys())
    num_heliostats = [helios.tilt[f].shape[0] for f in files]
    same_dust = np.zeros((len(files),len(files)))
    same_ext = [ [[] for n in range(num_heliostats[f])] for f in files]
                
    for ii,f in enumerate(files):
        for jj,g in enumerate(files):
            if len(D[f]) == len(D[g]):
                same_diameters =  np.all( D[f] == D[g])
            else:
                same_diameters = False

            if len(lam[f]) == len(lam[g]):
                same_lams =  np.all(lam[f] == sim_dat.source_wavelength[f])
            else:
                same_lams = False

            if len(intensities[f]) == len(intensities[g]):
                same_intensity =  np.all(intensities[f] == intensities[g])
            
            same_ref_ind =  (refractive_index[f] == refractive_index[g])
            same_dust[ii,jj] = same_diameters and same_lams \
                and same_intensity and same_ref_ind

    for ii,f in enumerate(files):
        for jj in range(num_heliostats[f]):
            for kk,g in enumerate(files):
                if same_dust[ii,kk]:
                    a = phia[f][jj]
                    idx = [(g,mm) for mm,pg in enumerate(phia[g]) if pg==a]
                    same_ext[ii][jj].extend(idx)
                
    return same_ext

def set_extinction_coefficients(destination_model,extinction_weights,file_inds):
    """
    Directly set extinction weights, e.g. from another model.

    This function sets the extinction weights directly. The input extinction_weights
    is an H-by-D numpy.array, where H is the number of heliostats and D is the number of 
    diameter bins. It primary use is to save time, since computation of the Mie 
    extinction weights can be time-consuming. 
    
    The required argument file_inds is a list, and the extinction weights of those 
    each of these files will be set to the np.array extinction_weights. Note that 
    the zeroth dimension of extinction_weights must be the same as the number of heliostats 
    in the destination file (determined by destination_model.helios.tilt). 
    
    """
    dm = destination_model
    ew = extinction_weights

    for f in file_inds:
        H = dm.helios.tilt[f].shape[0]
        D = ew.shape[1]
        if ew.shape[0] == H:
            dm.helios.extinction_weighting[f] = ew
        elif ew.shape[0] == 1:
            print(  f"Warning: ext_weights had only one heliostat. Broadcasting up to {H} heliostats "+
                    f"present in the destination_model.helios.tilt in file {f}.")
            dm.helios.extinction_weighting[f] = np.zeros((H,D))
            for h in range(H):
                dm.helios.extinction_weighting[f][h,:] = ew
        else:
             raise ValueError(  "Number of heliostats in extinction_weights (dim0) must either be one or the same as "+
                                "those in destination_model.helios.tilt")
    return dm

def get_training_data(d,file_start,time_to_remove_at_end):
    files = [f for f in os.listdir(d) if f.startswith(file_start)]

    # get training time intervals
    if np.isscalar(time_to_remove_at_end):
        time_to_remove_at_end = [time_to_remove_at_end]*len(files)
    training_intervals = []
    parse_date = lambda x: np.datetime64(f"{x[0:4]}-{x[4:6]}-{x[6::]}T00:00:00")
    for ii,f in enumerate(files):
        f = f.split(".")[0]
        dates = [parse_date(s) for s in f.split("_") if s.isnumeric()]
        assert len(dates)==2, "File name must contain start and end dates in YYYYMMDD format."
        s = min(dates)
        e = max(dates)

        e += np.timedelta64(1,'D') # since I'm appending midnight, need to use next day to get all data
        e -= np.timedelta64(time_to_remove_at_end[ii],'h') # leave specified testing time at the end (in hours)
        training_intervals.append(np.array([s,e]))

    training_intervals = np.stack(training_intervals).astype('datetime64[m]')

    # get mirror names in each file
    mirror_names = [ [] for f in files]
    for ii,f in enumerate(files):
        mirror_names[ii] = list(pd.read_excel(d+f,sheet_name="Reflectance_Average").columns[1::])

    # get mirror names that show up in all files
    common = []
    for a in mirror_names:
        for ele in a:
            if all([(ele in S) and (ele not in common) for S in mirror_names]):
                common.append(ele)

    files = [d+f for f in files]

    return files,training_intervals,mirror_names,common

def _parse_dust_str(dust_type):
    assert dust_type.startswith(("TSP","PM")), "dust_type must be PMX, PMX.X or TSP"
    if dust_type.startswith("TSP"):
        attr = "TSP"
    else:
        attr = dust_type[2::]
        if "." not in attr: # integer, e.g. PM20
            attr = f"PM{attr}"
        else: # decimal, e.g. PM2.5
            attr = "PM"+"_".join(attr.split('.'))
    return attr

def wind_rose(simulation_data,exp_idx):
    
    from windrose import WindroseAxes
    fig = plt.figure()
    wd = simulation_data.wind_direction[exp_idx]
    ws = simulation_data.wind_speed[exp_idx]
    wax = WindroseAxes.from_ax(fig=fig)
    wax.bar(wd,ws,normed=True)
    wax.set_legend()

    return fig,wax

def soiling_rates_summary(ref_data,sim_data,verbose=False):

    soiling_rates = ref_data.soiling_rate
    ave_data = ref_data.average

    # Determine starting index based on file name
    idx_start = 0
    if any("augusta".lower() in value.lower() for value in sim_data.file_name.values()):
        idx_start = 1  # For Port Augusta data
    if any("mildura".lower() in value.lower() for value in sim_data.file_name.values()):
        idx_start = 2  # For Mildura data

    # Initialize a dictionary to store soiling rates by campaign
    # soiling_rate_groups_by_campaign = {}
    results = []  # List to store results for DataFrame

    # Process each campaign
    for campaign_id, rates in soiling_rates.items():
        # Compute elapsed time
        elapsed_time = (ref_data.times[campaign_id][-1]-ref_data.times[campaign_id][0]) / np.timedelta64(1, 'D')

        # Extract campaign-specific tilts
        tilts = np.squeeze(np.unique(ref_data.tilts[campaign_id], axis=1))

        # Initialize a dictionary to store soiling rates for this campaign
        soiling_rate_groups = defaultdict(list)
        ave_groups = defaultdict(list)

        # Collect rates grouped by tilt, and corresponding reflectance values
        for tilt, rate, ave in zip(tilts[idx_start:], rates[idx_start:], ave_data[campaign_id].T[idx_start:]):
            soiling_rate_groups[tilt].append(rate)
            ave_groups[tilt].append(ave)

        # Compute the average soiling rate for each tilt
        campaign_averages_soiling = {
            tilt: np.mean(rates) for tilt, rates in soiling_rate_groups.items()
        }

        # Compute the initial and final ave values for each tilt
        for tilt, aves in ave_groups.items():
            initial_ave = np.mean([array[0] for array in aves]) if aves else np.nan
            final_ave = np.mean([array[-1] for array in aves]) if aves else np.nan
            tot_loss = initial_ave - final_ave
            avg_rate = campaign_averages_soiling.get(tilt, np.nan)
            results.append([campaign_id, elapsed_time, tilt, initial_ave, final_ave, tot_loss, avg_rate])

    # Create a DataFrame from the results
    df_ref_data = pd.DataFrame(results, columns=["Campaign", "Elapsed Time (days)", "Tilt", "Initial Reflectance", "Final Reflectance", "Total Loss", "Soiling Rate"])

# Display the results
    if verbose:
        print("\nAverage Soiling Rate and Initial/Final ave Values (By Campaign):")
        for campaign_id in soiling_rates:
            print(f"Campaign {campaign_id}:")
            # Filter results for the current campaign
            campaign_results = [res for res in results if res[0] == campaign_id]
            
            print("  Average Soiling Rates by Tilt:")
            for res in campaign_results:
                tilt = res[1]
                avg_rate = res[2]
                print(f"    Tilt {tilt}: {avg_rate:.2f}")
            
            print("  Initial and Final ave Values by Tilt:")
            for res in campaign_results:
                tilt = res[1]
                initial_ave = res[3]
                final_ave = res[4]
                print(f"    Tilt {tilt}: Initial ave = {initial_ave*100:.1f}%, Final ave = {final_ave*100:.1f}%")

    return df_ref_data

def loss_table_from_sim(sim_res,sim_data):
    table_data = []

    # Iterate over campaigns and tilts in the simulation data
    for (campaign_idx, tilt), y_data in sim_res.items():
        # Extract time data for the current campaign
        time_data = sim_data.time[campaign_idx]
        
        # # Ensure time_data and y_data lengths match
        if len(time_data) != len(y_data):
            raise ValueError(f"Time and Y data lengths do not match for campaign {campaign_idx}, tilt {tilt}.")
        
        # Calculate elapsed time in days
        elapsed_time = (time_data[time_data.index[-1]] - time_data[0]) / np.timedelta64(1, 'D')
        
        # Calculate initial value, final value, total loss, and average daily loss
        initial_value = y_data[0]
        final_value = y_data[-1]
        total_loss = initial_value - final_value
        avg_daily_loss = total_loss / elapsed_time if elapsed_time > 0 else np.nan
        
        # Append the data to the table
        table_data.append({
            "Campaign": f"Campaign {campaign_idx + 1}",
            "Elapsed Time (days)": elapsed_time,
            "Tilt": tilt,
            "Initial Value": initial_value,
            "Final Value": final_value,
            "Total Loss": total_loss,
            "Average Daily Loss": avg_daily_loss
        })

    # Convert to DataFrame for easier handling and export
    df_sim = pd.DataFrame(table_data)

    # Save to CSV
    # output_file = "loss_summary.csv"
    # df_sim.to_csv(output_file, index=False)

    # Display the DataFrame
    # print(df_sim)

    return df_sim

def loss_hel_table_from_sim(sim_res_hel,sim_data):  # provide simulated reflectance losses for each heliostats
    table_data = []

    # Iterate over campaigns and tilts in the simulation data
    for campaign_idx in range(len(sim_res_hel)):
        
        for hel, y_data in sim_res_hel[campaign_idx].items():
           
            # Calculate elapsed time in days
            elapsed_time = (y_data['Time'][-1]-y_data['Time'][0]).astype('timedelta64[s]')
            elapsed_time = pd.Timedelta(elapsed_time)/ np.timedelta64(1, 'D')
            
            # Calculate initial value, final value, total loss, and average daily loss
            initial_value = y_data['Reflectance'][0]
            final_value = y_data['Reflectance'][-1]
            total_loss = initial_value - final_value
            avg_daily_loss = total_loss / elapsed_time if elapsed_time > 0 else np.nan
            
            # Append the data to the table
            table_data.append({
                "Campaign": f"Campaign {campaign_idx + 1}",
                "Elapsed Time (days)": elapsed_time,
                "Heliostat": hel,
                "Initial Value": initial_value,
                "Final Value": final_value,
                "Total Loss": total_loss,
                "Average Daily Loss": avg_daily_loss
            })

    # Convert to DataFrame for easier handling and export
    df_sim_hel = pd.DataFrame(table_data)

    # Save to CSV
    # output_file = "loss_summary.csv"
    # df_sim.to_csv(output_file, index=False)

    # Display the DataFrame
    # print(df_sim)

    return df_sim_hel

# # EXTRACT VALUES FROM PLOT (NEED FIXING!)
def loss_table_from_plot(ax):
    table_data = []

    # Iterate over rows (tilts) and columns (campaigns)
    for campaign_idx, row in enumerate(ax.T):  # Transpose to iterate over columns (campaigns)
        for _ , subplot in enumerate(row):  # Iterate over individual subplots (tilts)
            # Check if the word "Tilt" is in the title of the subplot
            if "Tilt" in subplot.get_title():
                lines = subplot.get_lines()
                
                # Process only the first line (Mean)
                line = lines[0]  # The first line is the mean line
                x_data = line.get_xdata()
                y_data = line.get_ydata()
                
                # Calculate metrics
                initial_value = y_data[0]         # Initial value
                final_value = y_data[-1]          # Final value
                elapsed_time = x_data[-1] - x_data[0]  # Time difference
                
                # Ensure no division by zero
                if elapsed_time > 0:
                    avg_daily_loss = (initial_value - final_value) / elapsed_time
                else:
                    avg_daily_loss = np.nan  # Handle edge case for no elapsed time
                
                # Add to the table
                table_data.append({
                    "Campaign": f"Campaign {campaign_idx + 1}",
                    "Tilt": subplot.get_title().split(" ")[-1],  # Extract the tilt from the title
                    "Line Type": "Mean",
                    "Initial Value": initial_value,
                    "Final Value": final_value,
                    "Elapsed Time (days)": elapsed_time,
                    "Total Loss": initial_value - final_value,
                    "Average Daily Loss": avg_daily_loss
                })

    # Convert to DataFrame for easier handling and export
    df_plot = pd.DataFrame(table_data)

    # # Save to CSV
    # df_plot.to_csv(cm_save_file+"loss_summary.csv", index=False)

    # Display the DataFrame
    # print(df_plot)

    return df_plot

class DustDistribution():
    """
        This needs a docstring :(
    """
    def __init__(self,params=None,type=None):
        
        self.n_components = None
        self.sub_dists = []
        self.weights = []
        self.type = []
        self.units = []
        if params is not None:
            assert type.lower() in ["mass","number","area"], "Please supply a type (mass, number, area)"
            N = len(params)/3
            assert np.abs(N-np.floor(N))<np.finfo(float).eps, \
                "Please specify parameters of each component as a 1D numpy.array([weights,mu,sigma])."
            
            N = int(np.floor(N))
            w,mu,sig = params[0:N],params[N:2*N],params[2*N::]
            self.n_components = N
            self.sub_dists = [sps.norm(loc=mu[ii],scale=sig[ii]) for ii in range(N)]
            self.weights = w
            self.type = type
            self._set_units()

    def pdf(self,x):
        pdf = 0
        for ii in range(self.n_components):
            pdf += self.weights[ii]*self.sub_dists[ii].pdf(x)
        return pdf
    
    def cdf(self,x):
        cdf = 0
        for ii in range(self.n_components):
            cdf += self.weights[ii]*self.sub_dists[ii].cdf(x)
        return cdf
    
    def mean(self):
        m = 0
        sum_weights = sum(self.weights)
        for ii in range(self.n_components):
            m += self.weights[ii]/sum_weights*self.sub_dists[ii].mean()
        return m
    
    def icdf(self,p):
        fun = lambda x: self.cdf(x)-p
        res = spo.fsolve(fun,self.mean())
        return res
    
    def _sse(self,params,log_diameter_values,pm_values):

        N = len(params)/3
        assert np.abs(N-np.floor(N))<np.finfo(float).eps, \
            "Please specify parameters of each component as a 1D numpy.array([weights,mu,sigma])."
        N = int(np.floor(N))
        w,mu,sig = params[0:N],params[N:2*N],params[2*N::]
        self.n_components = N
        self.sub_dists = [sps.norm(loc=mu[ii],scale=sig[ii]) for ii in range(N)]
        self.weights = w

        return np.sum( (self.cdf(log_diameter_values)-pm_values)**2 )
    
    def fit(self,params0,log_diameter_values,pm_values,tol=1e-3):

        N = len(params0)/3
        assert np.abs(N-np.floor(N))<np.finfo(float).eps, \
            "Please specify parameters of each component as a 1D numpy.array([weights,mu,sigma])."
        N = int(np.floor(N))

        fun = lambda x: self._sse(x,log_diameter_values,pm_values)
        
        # construct bounds
        lower_bound_w = [0]*N
        lower_bound_mu = [-np.inf]*N
        lower_bound_sig = [0+tol]*N
        lb = lower_bound_w + lower_bound_mu + lower_bound_sig # join lists
        ub = [np.inf]*len(lb)

        bnds = spo.Bounds(lb=lb,ub=ub,keep_feasible=True)
        res = spo.minimize(fun,params0,bounds=bnds,tol=1e-8)

        params = res.x
        N = int(np.floor(len(params)/3))
        w,mu,sig = params[0:N],params[N:2*N],params[2*N::]
        self.n_components = N
        self.sub_dists = [sps.norm(loc=mu[ii],scale=sig[ii]) for ii in range(N)]
        self.weights = w
        self.type = "mass"
        self._set_units()

        return res

    def _set_units(self):
        if self.type.lower() == "mass":
            self.units = r"$\frac{\mu g \cdot m^{{-3}}}{d(\log(D))}$"
        elif self.type.lower() == "number":
            self.units = r"$\frac{cm^{{-3}} ] }{d(\log(D))}$"
        elif self.type.lower() == "area":
            self.units = r"$\frac{ m^2 \cdot m^{{-3}}}{d(\log(D))}$"

    def convert_to_number(self,rho=None):

        if self.type.lower() == "number":
            print("Type is already ""number"" ")
        elif self.type.lower()=="mass":
            assert isinstance(rho,float), "Particle density must be a scalar float."
            new_subs = []
            new_weights = []
            for ii in range(self.n_components):
                Mi = self.weights[ii]
                mbari = self.sub_dists[ii].mean()
                si = self.sub_dists[ii].std()

                mi = mbari-3*si**2/np.log10(np.e)
                b = 2*mi+6*si**2/np.log10(np.e)
                Ni = 6*Mi/np.pi/rho * np.exp((mi**2-0.25*b**2)/2/si**2)

                new_weights.append(Ni*1e3)
                ns = sps.norm(loc=mi,scale=si)
                new_subs.append(ns)
            
            self.weights = new_weights
            self.sub_dists = new_subs
            self.type = "number"
            self._set_units()
        elif self.type.lower()=="area":
            new_subs = []
            new_weights = []
            for ii in range(self.n_components):
                Ai = self.weights[ii]
                mbari = self.sub_dists[ii].mean()
                si = self.sub_dists[ii].std()

                mi = mbari-2*si**2/np.log10(np.e)
                Ni = Ai/np.pi*4 * np.exp((mi**2-(mi-si**2/np.log10(np.e))**2)/2/si**2)

                new_weights.append(Ni*1e6)
                ns = sps.norm(loc=mi,scale=si)
                new_subs.append(ns)
            
            self.weights = new_weights
            self.sub_dists = new_subs
            self.type = "number"
            self._set_units()

    def convert_to_mass(self,rho):
        assert isinstance(rho,float), "Particle density must be a scalar float."
        if self.type.lower() == "mass":
            print("Type is already ""mass"" ")
        elif self.type.lower() == "number":
            new_subs = []
            new_weights = []
            for ii in range(self.n_components):
                Ni = self.weights[ii]
                mi = self.sub_dists[ii].mean()
                si = self.sub_dists[ii].std()

                b = 2*mi+6*si**2/np.log10(np.e)
                mbari = b/2.0
                Mi = Ni*np.pi*rho/6 * np.exp(-(mi**2-0.25*b**2)/2/si**2)

                new_weights.append(Mi*1e-3)
                ns = sps.norm(loc=mbari,scale=si)
                new_subs.append(ns)
            
            self.weights = new_weights
            self.sub_dists = new_subs
            self.type = "mass"
            self._set_units()
        elif self.type.lower()=="area":
            print("Convert to number first. ")

    def convert_to_area(self,rho=None):

        if self.type.lower() == "area":
            print("Type is already ""area"" ")
        elif self.type.lower() == "number":
            new_subs = []
            new_weights = []
            for ii in range(self.n_components):
                Ni = self.weights[ii]
                mi = self.sub_dists[ii].mean()
                si = self.sub_dists[ii].std()

                mbari = mi+2*si**2/np.log10(np.e)
                Ai = Ni*np.pi/4*np.exp(-(mi**2-(mi+si**2/np.log10(np.e))**2)/2/si**2)

                new_weights.append(Ai*1e-6)
                ns = sps.norm(loc=mbari,scale=si)
                new_subs.append(ns)
            
            self.weights = new_weights
            self.sub_dists = new_subs
            self.type = "area"
            self._set_units()
        elif self.type.lower() == "mass":
            if rho is None:
                raise ValueError("Rho cannot be None to convert from mass")
            
            assert isinstance(rho,float), "Particle density must be a scalar float."
            self.convert_to_number(rho)
            self.convert_to_area()

    def write_to_file(self,file_name,sheet_name,kind='number',rho=None,verbose=True):
        _print_if("Writing dust distribution to file "+file_name,verbose)
        
        # ensure kind is correct
        if kind.lower() == 'number':
            self.convert_to_number(rho)
        elif kind.lower(rho) == 'mass':
            self.convert_to_mass(rho)
        elif kind.lower(rho) == 'area':
            self.convert_to_area(rho)
        else:
            raise ValueError("kind not recognized.")

        # convert to strings and join with ";" delimiter
        weight_str = [str(s)+";" for s in self.weights]
        mu_str = [str(10**(self.sub_dists[ii].mean())) for ii in range(self.n_components)]
        sig_str = [str(10**(self.sub_dists[ii].std())) for ii in range(self.n_components)]
        weight_str = "".join(weight_str)[0:-1]
        mu_str = "".join(mu_str)[0:-1]
        sig_str = "".join(sig_str)[0:-1]

        # write data
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
        for cell in ws['A']:
            if cell.value == "N_size":
                ws.cell(row=cell.row, column=2).value = self.n_components
                ws.cell(row=cell.row, column=4).value = ""
            elif cell.value == "Nd":
                ws.cell(row=cell.row, column=2).value = weight_str
                ws.cell(row=cell.row, column=4).value = ""
            elif cell.value == "mu":
                ws.cell(row=cell.row, column=2).value = mu_str
                ws.cell(row=cell.row, column=4).value = ""
            elif cell.value == "sigma":
                ws.cell(row=cell.row, column=2).value = sig_str
                ws.cell(row=cell.row, column=4).value = ""
        
        wb.save(filename=file_name)
        wb.close()


